import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

"""def marcarDuplicadosEnExcel(nombre_archivo, duplicados):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    rojo = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for i, es_duplicado in enumerate(duplicados, start=2):
        if es_duplicado:
            ws[f"Y{i}"].fill = rojo
    wb.save(nombre_archivo)"""

def borrarORG(df, iata):
    if iata == "crd".upper():
            df.loc[df["Nombre Solicitante"] == "ORG COURIER ARG", "Ruta Virtual"] = 700
    elif iata == "luq".upper():
        df.loc[df["Nombre Solicitante"] == "ORG COURIER ARG", "Ruta Virtual"] = 700
    elif iata == "irj".upper():
        pass
    else:
        print("error")
    return df

def borrarMHTML(): #si encuentra archivos mhtml preguntar si lo queres borrar
    while True:
        pregunta = input("queres borrar los archivos MHTML?: S/N ").lower()
        if pregunta == "s":
            encontrar = glob.glob("*MHTML")
            if not encontrar:
                print("no hay archivos MHTML")
            else:
                for archivo in encontrar:
                    os.remove(archivo)
            break
        if pregunta == "n":
            print("no se borraron archivos")
            break
        else:
            print("Ingresa 's' o 'n'")

def ejecutarExcelFinalizado(iata):
    os.startfile(f"archivoUnificado{iata}.xlsx")

def manipularDatos(df, iata):
    # filtrar columnas
    df = df [
        (df["Motivo Descripción"] != "Retirado") & 
        (df["Motivo Descripción"] != "Entregado") &
        (df["Destino"] == iata)
    ].copy()

    # agregar valores
    df.loc[df["Peso del objeto"] >= 200, "Ruta Virtual"] = 503
    df.loc[df["Volumen"] >= 1.5, "Ruta Virtual"] = 503

    #limpiar columnas
    df["Distrito Destino"] = ""
    df["Provincia"] = ""

    duplicados = df.duplicated(subset= "Nro. identificación pieza según cliente", keep = False)
    return df, duplicados

def main():
    while True:
        iata = input("Ingresa el codigo IATA: ").upper()
        if len(iata) == 3:
            break
        else:
            print("error intente nuevametne")

    if iata == "crd".upper(): #listo
        encontrar = glob.glob("*xlsx")
        lista = []
        for archivo in encontrar:
            leer = pd.read_excel(archivo)
            lista.append(leer)
        
        df = pd.concat(lista, ignore_index=True)
        df, duplicados  = manipularDatos(df, iata)
        df = borrarORG(df, iata)  
        
        df.to_excel(f"archivoUnificado{iata}.xlsx", index=False)
        #marcarDuplicadosEnExcel(f"archivoUnificado{iata}.xlsx", duplicados)
        borrarMHTML()
        ejecutarExcelFinalizado(iata)

    elif iata == "luq".upper(): # listo
        encontrar = glob.glob("*xlsx")
        lista = []
        for archivo in encontrar:
            leer = pd.read_excel(archivo)
            lista.append(leer)
        
        df = pd.concat(lista, ignore_index=True)
        df, duplicados  = manipularDatos(df, iata)
        df = borrarORG(df, iata)  
        
        df.to_excel(f"archivoUnificado{iata}.xlsx", index=False)
        #marcarDuplicadosEnExcel(f"archivoUnificado{iata}.xlsx", duplicados)
        borrarMHTML()
        ejecutarExcelFinalizado(iata)

    elif iata == "irj".upper():
        encontrar = glob.glob("*xlsx")
        lista = []
        for archivo in encontrar:
            leer = pd.read_excel(archivo)
            lista.append(leer)
        df = pd.concat(lista, ignore_index=True)
        #solo vaciar columnas 
        df["Distrito Destino"] = ""
        df["Provincia"] = ""
        df.to_excel(f"archivoUnificado{iata}.xlsx", index=False)
        borrarMHTML()
        ejecutarExcelFinalizado(iata)
    
    elif iata == "fma".upper():
        encontrar = glob.glob("*xlsx")
        lista = []
        for archivo in encontrar:
            leer = pd.read_excel(archivo)
            lista.append(leer)
        df = pd.concat(lista, ignore_index=True)
        df = manipularDatos(df, iata)
        

    else:
        print("error")
if __name__ == "__main__": 
    main()